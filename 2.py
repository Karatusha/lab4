import openpyxl
from openpyxl.styles import Font
import csv
from datetime import datetime, date, timedelta

try:
    workbook = openpyxl.Workbook()
    sheets = {
        "all": workbook.active,
        "younger_18": workbook.create_sheet("younger_18"),
        "18-45": workbook.create_sheet("18-45"),
        "45-70": workbook.create_sheet("45-70"),
        "older_70": workbook.create_sheet("older_70")
    }

    for sheet_name in sheets:
       sheets[sheet_name].append(['№', 'Прізвище', 'Ім\'я', 'По батькові', 'Дата народження','Вік'])


    try:
        with open('employees.csv', 'r', newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            index = 1
            for row in reader:
                birth_date_str = row['Дата народження']
                birth_date = datetime.strptime(birth_date_str, '%Y-%m-%d').date()
                age = (date.today() - birth_date) // timedelta(days=365.2425)  # More accurate age calculation

                row_data = [index, row['Прізвище'], row['Ім\'я'], row['По батькові'], row['Дата народження'], age]

                sheets["all"].append(row_data)

                if age < 18:
                    sheets["younger_18"].append(row_data)
                elif 18 <= age <= 45:
                    sheets["18-45"].append(row_data)
                elif 45 < age <= 70:
                    sheets["45-70"].append(row_data)
                else:
                    sheets["older_70"].append(row_data)

                index += 1


        workbook.save('employees.xlsx')
        print("Ok")

    except FileNotFoundError:
        print("Повідомлення про відсутність, або проблеми при відкритті файлу CSV.")

except Exception as e:
    print(f"Повідомлення про неможливість створення XLSX файлу: {e}")